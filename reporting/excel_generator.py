"""Excel report generator.

Produces a multi-sheet Excel workbook:
  - Sheet 1: All Tickets (one row per ticket, passengers grouped)
  - Sheet 2: Passenger Journey Summary (one row per passenger-journey)
"""

import logging
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from models.ticket import Ticket
from processing.journey_aggregator import JourneyEntry
from utils.constants import SHEET_ALL_TICKETS, SHEET_PASSENGER_SUMMARY

logger = logging.getLogger("irctc_analyzer")

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
PASSENGER_HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")


def generate_excel(
    tickets: list[Ticket],
    passenger_journeys: dict[str, list[JourneyEntry]],
    output_path: Path,
) -> None:
    """Generate the multi-sheet Excel report.

    Args:
        tickets: Chronologically sorted ticket list.
        passenger_journeys: Per-passenger journey mapping.
        output_path: Full path for the output .xlsx file.
    """
    logger.info("Generating Excel report: %s", output_path.name)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _write_all_tickets_sheet(writer, tickets)
        _write_passenger_summary_sheet(writer, passenger_journeys)

    # Post-process styling
    _apply_styling(output_path)

    logger.info("[OK] Excel report saved: %s", output_path)


# ======================================================================
# Sheet 1: All Tickets
# ======================================================================


def _write_all_tickets_sheet(
    writer: pd.ExcelWriter,
    tickets: list[Ticket],
) -> None:
    """Write the All Tickets sheet with one row per ticket."""
    rows: list[dict] = []

    for ticket in tickets:
        rows.append(
            {
                "Passenger Names": ticket.passenger_names,
                "PNR": ticket.pnr,
                "Train Number": ticket.train_number,
                "Train Name": ticket.train_name,
                "From": ticket.from_station,
                "To": ticket.to_station,
                "Departure": ticket.departure_display,
                "Arrival": ticket.arrival_display,
                "Seat Numbers": ticket.seat_numbers,
                "Coach": ticket.coaches,
                "Booking Status": ticket.booking_statuses,
                "Current Status": ticket.current_statuses,
                "Class": ticket.travel_class,
                "Quota": ticket.quota,
                "Distance (km)": ticket.distance,
                "Ticket Fare": ticket.ticket_fare,
                "Convenience Fee": ticket.convenience_fee,
                "Insurance Fee": ticket.insurance_fee,
                "Total Fare": ticket.total_fare,
                "Booking Date": ticket.booking_display,
                "Transaction ID": ticket.transaction_id,
                "Invoice Number": ticket.invoice_number,
                "GSTIN": ticket.gstin,
                "SAC Code": ticket.sac_code,
                "File Name": ticket.file_name,
            }
        )

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name=SHEET_ALL_TICKETS, index=False)


# ======================================================================
# Sheet 2: Passenger Journey Summary
# ======================================================================


def _write_passenger_summary_sheet(
    writer: pd.ExcelWriter,
    passenger_journeys: dict[str, list[JourneyEntry]],
) -> None:
    """Write the Passenger Journey Summary sheet."""
    rows: list[dict] = []

    for name, journeys in passenger_journeys.items():
        for journey in journeys:
            rows.append(
                {
                    "Passenger Name": name,
                    "Train Number": journey.train_number,
                    "Train Name": journey.train_name,
                    "PNR": journey.pnr,
                    "Departure": journey.departure_display,
                    "From": journey.from_station,
                    "To": journey.to_station,
                    "Seat (Status)": journey.seat_info,
                }
            )

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name=SHEET_PASSENGER_SUMMARY, index=False)


# ======================================================================
# Post-processing styles
# ======================================================================


def _apply_styling(output_path: Path) -> None:
    """Apply professional formatting to the generated workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(output_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        fill = HEADER_FILL if sheet_name == SHEET_ALL_TICKETS else PASSENGER_HEADER_FILL

        # Style header row
        for cell in ws[1]:
            cell.fill = fill
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

        # Auto-width columns and alternating row colors
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)

            for row_idx, cell in enumerate(col_cells, 1):
                # Calculate width
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

                # Data row styling
                if row_idx > 1:
                    cell.alignment = Alignment(
                        vertical="center", wrap_text=True
                    )
                    cell.border = THIN_BORDER

                    # Alternating row colors
                    if row_idx % 2 == 0:
                        cell.fill = ALT_ROW_FILL

            # Set column width (capped)
            adjusted_width = min(max_length + 4, 40)
            ws.column_dimensions[col_letter].width = max(adjusted_width, 12)

        # Freeze top row
        ws.freeze_panes = "A2"

    wb.save(output_path)
